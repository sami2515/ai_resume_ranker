"""
Export endpoint (project documentation Section 4.4, FR-08)
Produces executive-grade Excel (.xlsx) and PDF (.pdf) candidate ranking reports.
"""

from __future__ import annotations
import io
import os
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file, g

from models import JobDescriptionModel, RecruiterFeedback, MatchResult
from services import get_ranked_results
from auth import require_auth
from extensions import db
from .jobs import _get_owned_job

export_bp = Blueprint("export", __name__, url_prefix="/api/export")

COLUMNS = [
    "Rank", "Candidate", "Score", "Keyword", "Semantic",
    "Confidence", "Matched Skills", "Missing Skills", "Exp. (yrs)", "Decision"
]

BRAND_DARK  = "#090D16"
BRAND_BLUE  = "#2563EB"
BRAND_SLATE = "#1E293B"
BRAND_LIGHT = "#F8FAFC"
BRAND_GREEN = "#10B981"
BRAND_SKY   = "#0EA5E9"
BRAND_AMBER = "#F59E0B"
BRAND_RED   = "#F43F5E"


def _get_tier_meta(conf_str: str) -> dict:
    """Single source of truth for 4-tier match confidence colors and labels in exports."""
    s = str(conf_str).lower()
    if "high" in s:
        return {"short": "High Match", "color": "#10B981", "bg": "#D1FAE5", "text": "#065F46"}
    elif "moderate" in s:
        return {"short": "Moderate Match", "color": "#0EA5E9", "bg": "#E0F2FE", "text": "#075985"}
    elif "partial" in s or "weak" in s:
        return {"short": "Weak Match", "color": "#F59E0B", "bg": "#FEF3C7", "text": "#92400E"}
    else:
        return {"short": "No Match", "color": "#F43F5E", "bg": "#FFE4E6", "text": "#991B1B"}


def _rows(job_id: int, decisions: dict):
    for i, row in enumerate(get_ranked_results(job_id), 1):
        c = row.candidate
        decision = decisions.get(row.id, "—")
        yield [
            i,
            c.full_name or c.resume_filename,
            f"{row.composite_score:.1f}",
            f"{row.keyword_score:.1f}",
            f"{row.semantic_score:.1f}",
            row.confidence,
            ", ".join(row.matched_skills or []) or "—",
            ", ".join(row.missing_skills or []) or "—",
            f"{c.experience_years or 0:.1f}",
            decision.capitalize() if decision != "—" else "—",
        ]


def _get_decisions(job_id: int) -> dict:
    """Return {match_result_id: decision} for the requesting recruiter."""
    match_ids = [r.id for r in MatchResult.query.filter_by(jd_id=job_id).all()]
    if not match_ids:
        return {}
    rows = (
        RecruiterFeedback.query
        .filter(
            RecruiterFeedback.match_result_id.in_(match_ids),
            RecruiterFeedback.recruiter_id == g.recruiter_id,
        )
        .order_by(RecruiterFeedback.created_at.asc())
        .all()
    )
    return {r.match_result_id: r.decision for r in rows}


@export_bp.get("/<int:job_id>")
@require_auth
def export_results(job_id: int):
    jd, error = _get_owned_job(job_id)
    if error:
        return error

    fmt = request.args.get("format", "excel").lower()
    decisions = _get_decisions(job_id)
    rows = list(_rows(job_id, decisions))

    if fmt in ("excel", "xlsx"):
        return _export_excel(jd, rows)
    if fmt == "pdf":
        return _export_pdf(jd, rows, decisions)
    return jsonify(error="format must be 'excel' or 'pdf'."), 400


def _export_excel(jd: JobDescriptionModel, rows: list[list]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"AI ResumeRanker — {jd.title}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="2563EB")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sub-header row
    ws.merge_cells("A2:J2")
    sub_cell = ws["A2"]
    sub_cell.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Category: {jd.category or 'General'}  |  "
        f"Total Evaluated: {len(rows)} candidates"
    )
    sub_cell.font = Font(italic=True, size=9, color="64748B")
    sub_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Column headers
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 22

    # Data rows
    fills = {
        "High Match": PatternFill("solid", fgColor="D1FAE5"),
        "Moderate Match": PatternFill("solid", fgColor="E0F2FE"),
        "Weak Match": PatternFill("solid", fgColor="FEF3C7"),
        "No Match": PatternFill("solid", fgColor="FFE4E6"),
    }
    font_colors = {
        "High Match": "065F46",
        "Moderate Match": "075985",
        "Weak Match": "92400E",
        "No Match": "991B1B",
    }
    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    for row_idx, row in enumerate(rows, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 3, 4, 5, 9, 10) else "left",
                vertical="center"
            )
            # Confidence color coding
            if col_idx == 6:
                tier = _get_tier_meta(str(value))
                short_name = tier["short"]
                cell.value = short_name
                cell.fill = fills.get(short_name, fills["No Match"])
                cell.font = Font(color=font_colors.get(short_name, "991B1B"), bold=True, size=9)
            elif row_idx % 2 == 1:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    col_widths = [6, 24, 8, 9, 9, 20, 35, 35, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"AI_ResumeRanker_{jd.title.replace(' ','_')}_job{jd.id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _export_pdf(jd: JobDescriptionModel, rows: list[list], decisions: dict):
    """Generates an executive, beautifully styled PDF candidate evaluation report."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfgen import canvas as pdfcanvas

    buf = io.BytesIO()
    page_w, page_h = landscape(A4)  # 841.89 x 595.27 points (29.7cm x 21.0cm)

    total_candidates = len(rows)
    high_count = sum(1 for r in rows if _get_tier_meta(r[5])["short"] == "High Match")
    mod_count  = sum(1 for r in rows if _get_tier_meta(r[5])["short"] == "Moderate Match")
    weak_count = sum(1 for r in rows if _get_tier_meta(r[5])["short"] == "Weak Match")
    no_count   = sum(1 for r in rows if _get_tier_meta(r[5])["short"] == "No Match")

    hired_count    = sum(1 for r in rows if str(r[9]).lower() == "hired")
    rejected_count = sum(1 for r in rows if str(r[9]).lower() == "rejected")

    # ── Executive Branded Canvas ─────────────────────────────────────────────
    class BrandedPageCanvas(pdfcanvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_branded_chrome(num_pages)
                pdfcanvas.Canvas.showPage(self)
            pdfcanvas.Canvas.save(self)

        def _draw_branded_chrome(self, total_pages):
            # Top Banner Header Bar
            self.setFillColor(colors.HexColor("#0F172A"))
            self.rect(0, page_h - 36, page_w, 36, fill=1, stroke=0)

            # Accent Underline
            self.setFillColor(colors.HexColor("#2563EB"))
            self.rect(0, page_h - 38, page_w, 2, fill=1, stroke=0)

            # Logo Badge Icon (Blue Square + Vector Star Target)
            self.setFillColor(colors.HexColor("#2563EB"))
            self.roundRect(18, page_h - 29, 22, 22, 5, fill=1, stroke=0)

            # Vector emblem inside logo badge
            self.setStrokeColor(colors.white)
            self.setLineWidth(1.4)
            self.circle(29, page_h - 18, 5, fill=0, stroke=1)
            self.setFillColor(colors.white)
            self.circle(29, page_h - 18, 2, fill=1, stroke=0)

            # Brand Title
            self.setFillColor(colors.white)
            self.setFont("Helvetica-Bold", 11)
            self.drawString(46, page_h - 23, "AI RESUMERANKER")

            # Top Right Job Title
            self.setFillColor(colors.HexColor("#94A3B8"))
            self.setFont("Helvetica-Bold", 9)
            self.drawRightString(page_w - 18, page_h - 22, f"{jd.title}  •  Category: {jd.category or 'General'}")

            # Bottom Footer Bar
            self.setFillColor(colors.HexColor("#1E293B"))
            self.rect(0, 0, page_w, 20, fill=1, stroke=0)

            self.setFillColor(colors.HexColor("#94A3B8"))
            self.setFont("Helvetica", 8)
            self.drawString(18, 6, f"Generated: {datetime.now().strftime('%b %d, %Y %H:%M')}  |  Confidential Recruiter Audit Report")

            self.setFont("Helvetica-Bold", 8)
            self.drawRightString(page_w - 18, 6, f"Page {self._pageNumber} of {total_pages}  |  Total Candidates: {total_candidates}")

    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=46,
        bottomMargin=26,
        leftMargin=18,
        rightMargin=18,
        title=f"AI ResumeRanker Report - {jd.title}",
        author="AI ResumeRanker System",
    )

    styles = getSampleStyleSheet()

    # Custom Paragraph Styles
    heading_style = ParagraphStyle(
        "ExecHeading",
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=16,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "ExecSub",
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "RRCell",
        fontName="Helvetica",
        fontSize=6.5,
        leading=8,
        textColor=colors.HexColor("#1E293B"),
    )
    cell_bold_center = ParagraphStyle(
        "RRCellBoldCenter",
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8.5,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0F172A"),
    )
    skill_cell_style = ParagraphStyle(
        "RRSkillCell",
        fontName="Helvetica",
        fontSize=6,
        leading=7.5,
        textColor=colors.HexColor("#334155"),
    )
    missing_skill_style = ParagraphStyle(
        "RRMissingSkillCell",
        fontName="Helvetica",
        fontSize=6,
        leading=7.5,
        textColor=colors.HexColor("#64748B"),
    )

    elements = []

    # ── Executive Overview Card ─────────────────────────────────────────────
    card_data = [
        [
            Paragraph("<b>EXECUTIVE EVALUATION SUMMARY</b>", ParagraphStyle("CardH", fontName="Helvetica-Bold", fontSize=9, textColor=colors.HexColor("#0F172A"))),
            Paragraph(f"Position: <b>{jd.title}</b>", ParagraphStyle("CardVal", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#334155"))),
            Paragraph(f"Category: <b>{jd.category or 'General'}</b>", ParagraphStyle("CardVal", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#334155"))),
            Paragraph(f"Evaluated: <b>{total_candidates} Candidates</b>", ParagraphStyle("CardVal", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#334155"))),
        ],
        [
            Paragraph(
                f"<font color='#065F46'><b>High Match: {high_count}</b></font> &nbsp;•&nbsp; "
                f"<font color='#075985'><b>Moderate Match: {mod_count}</b></font> &nbsp;•&nbsp; "
                f"<font color='#92400E'><b>Weak Match: {weak_count}</b></font> &nbsp;•&nbsp; "
                f"<font color='#991B1B'><b>No Match: {no_count}</b></font>",
                ParagraphStyle("CardTiers", fontName="Helvetica", fontSize=7.5, leading=9)
            ),
            Paragraph(
                f"Decisions: <font color='#10B981'><b>{hired_count} Hired</b></font> &nbsp;|&nbsp; "
                f"<font color='#F43F5E'><b>{rejected_count} Rejected</b></font>",
                ParagraphStyle("CardDec", fontName="Helvetica", fontSize=7.5, leading=9)
            ),
            Paragraph(
                f"Evaluation Date: <b>{datetime.now().strftime('%b %d, %Y')}</b>",
                ParagraphStyle("CardDate", fontName="Helvetica", fontSize=7.5, leading=9, textColor=colors.HexColor("#64748B"))
            ),
            Paragraph(
                "Status: <b>Verified by SBERT Pipeline</b>",
                ParagraphStyle("CardStatus", fontName="Helvetica-Bold", fontSize=7.5, leading=9, textColor=colors.HexColor("#2563EB"))
            ),
        ]
    ]

    card_table = Table(card_data, colWidths=[7.0*cm, 6.5*cm, 6.5*cm, 6.5*cm])
    card_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#E2E8F0")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#F1F5F9")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(card_table)
    elements.append(Spacer(1, 8))

    # ── Candidate Rankings Table ─────────────────────────────────────────────
    header_labels = ["#", "Candidate Name", "Score", "KW", "SEM", "Confidence Tier", "Matched Skills", "Missing Skills", "Exp", "Decision"]
    table_data = [[Paragraph(f"<b>{h}</b>", ParagraphStyle("HeaderCell", fontName="Helvetica-Bold", fontSize=7.5, textColor=colors.white, alignment=TA_CENTER)) for h in header_labels]]

    for row in rows:
        rank_num = row[0]
        name = str(row[1])
        composite = str(row[2])
        kw = str(row[3])
        sem = str(row[4])
        raw_conf = str(row[5])
        matched_str = str(row[6])
        missing_str = str(row[7])
        exp_str = str(row[8])
        decision_str = str(row[9])

        tier = _get_tier_meta(raw_conf)
        conf_short = tier["short"]
        conf_text_color = tier["text"]
        conf_bg = tier["bg"]

        if decision_str.lower() == "hired":
            dec_html = f"<font color='#10B981'><b>Hired</b></font>"
        elif decision_str.lower() == "rejected":
            dec_html = f"<font color='#F43F5E'><b>Rejected</b></font>"
        else:
            dec_html = f"<font color='#64748B'>—</font>"

        # Formatting skill paragraphs cleanly
        matched_p = Paragraph(matched_str, skill_cell_style)
        missing_p = Paragraph(missing_str, missing_skill_style)

        table_data.append([
            Paragraph(f"<b>#{rank_num}</b>", cell_bold_center),
            Paragraph(f"<b>{name}</b>", cell_style),
            Paragraph(f"<b>{composite}</b>", ParagraphStyle("ScoreC", fontName="Helvetica-Bold", fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor("#2563EB"))),
            Paragraph(kw, ParagraphStyle("KWC", fontName="Helvetica", fontSize=6.5, alignment=TA_CENTER)),
            Paragraph(sem, ParagraphStyle("SEMC", fontName="Helvetica", fontSize=6.5, alignment=TA_CENTER)),
            Paragraph(f"<font color='{conf_text_color}'><b>{conf_short}</b></font>", ParagraphStyle("ConfC", fontName="Helvetica-Bold", fontSize=7, alignment=TA_CENTER)),
            matched_p,
            missing_p,
            Paragraph(f"{exp_str} y", ParagraphStyle("ExpC", fontName="Helvetica", fontSize=6.5, alignment=TA_CENTER)),
            Paragraph(dec_html, ParagraphStyle("DecC", fontName="Helvetica-Bold", fontSize=7, alignment=TA_CENTER)),
        ])

    # Printable table width sum = 26.5 cm (Fits landscape A4 with 1.6cm left/right margins)
    col_widths_pdf = [1.0*cm, 4.5*cm, 1.4*cm, 1.2*cm, 1.2*cm, 2.8*cm, 6.2*cm, 5.2*cm, 1.2*cm, 1.8*cm]

    t = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    
    t_styles = [
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1E293B")),
        ("ALIGN",       (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",(0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]

    # Add custom tier badge background highlights per row
    for r_idx, row in enumerate(rows, 1):
        tier = _get_tier_meta(row[5])
        badge_bg_color = colors.HexColor(tier["bg"])
        t_styles.append(("BACKGROUND", (5, r_idx), (5, r_idx), badge_bg_color))
        if r_idx == 1:
            # Gold highlight for Rank 1
            t_styles.append(("BACKGROUND", (0, r_idx), (0, r_idx), colors.HexColor("#FEF3C7")))

    t.setStyle(TableStyle(t_styles))
    elements.append(t)

    doc.build(elements, canvasmaker=BrandedPageCanvas)
    buf.seek(0)

    safe_title = "".join(c for c in jd.title if c.isalnum() or c in (" ", "_", "-")).replace(" ", "_")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"AI_ResumeRanker_{safe_title}_Report.pdf",
        mimetype="application/pdf",
    )
